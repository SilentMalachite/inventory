from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Iterable, List, Tuple

from openpyxl import Workbook, load_workbook

from .models import Item


CSV_HEADERS = ["sku", "name", "category", "unit", "min_stock"]


def items_to_csv(items: Iterable[Item], encoding: str = "utf-8-sig") -> bytes:
    """Serialize items to CSV. Default encoding is utf-8-sig (BOM) for Windows Excel.
    Supported encodings: "utf-8", "utf-8-sig", "cp932" (Shift_JIS).
    """
    sio = io.StringIO(newline="")
    writer = csv.writer(sio)
    writer.writerow(CSV_HEADERS)
    for it in items:
        writer.writerow([
            it.sku,
            it.name,
            it.category or "",
            it.unit,
            it.min_stock,
        ])
    text = sio.getvalue()
    return text.encode(encoding)


def items_to_xlsx(items: Iterable[Item]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append(CSV_HEADERS)
    for it in items:
        ws.append([it.sku, it.name, it.category or "", it.unit, it.min_stock])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def dicts_to_csv(headers: list[str], rows: list[dict], encoding: str = "utf-8-sig") -> bytes:
    sio = io.StringIO(newline="")
    writer = csv.writer(sio)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h, "") for h in headers])
    return sio.getvalue().encode(encoding)


def parse_items_csv(data: bytes, encoding: str | None = None) -> List[dict]:
    """Parse CSV bytes into list of dicts. If encoding is None, try utf-8-sig then cp932.
    Returns list of dicts with keys CSV_HEADERS, min_stock as int.
    """
    errors: List[str] = []
    tried: List[str] = []
    for enc in ([encoding] if encoding else ["utf-8-sig", "cp932", "utf-8"]):
        if enc is None:
            continue
        tried.append(enc)
        try:
            text = data.decode(enc)
            break
        except Exception:
            text = None  # type: ignore
    else:
        raise ValueError(f"CSVのエンコード判定に失敗しました: {tried}")

    out: List[dict] = []
    reader = csv.DictReader(io.StringIO(text))  # type: ignore[arg-type]
    missing = [h for h in CSV_HEADERS if h not in reader.fieldnames[: len(CSV_HEADERS)] if reader.fieldnames]
    if missing:
        raise ValueError(f"CSVヘッダーが不足しています: {missing}")
    for row in reader:
        try:
            rec = {
                "sku": (row.get("sku") or "").strip(),
                "name": (row.get("name") or "").strip(),
                "category": (row.get("category") or None) or None,
                "unit": (row.get("unit") or "pcs").strip() or "pcs",
                "min_stock": int((row.get("min_stock") or 0) or 0),
            }
        except Exception as e:
            raise ValueError(f"CSVの行を解析できません: {row}") from e
        out.append(rec)
    return out


def parse_items_xlsx(data: bytes) -> List[dict]:
    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {name: headers.index(name) if name in headers else -1 for name in CSV_HEADERS}
    missing = [k for k, i in idx.items() if i < 0]
    if missing:
        raise ValueError(f"Excelのヘッダーが不足しています: {missing}")
    out: List[dict] = []
    for r in rows[1:]:
        rec = {
            "sku": (r[idx["sku"]] or "").strip(),
            "name": (r[idx["name"]] or "").strip(),
            "category": (r[idx["category"]] or None),
            "unit": (r[idx["unit"]] or "pcs").strip() or "pcs",
            "min_stock": int(r[idx["min_stock"]] or 0),
        }
        out.append(rec)
    return out
